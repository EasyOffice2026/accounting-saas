from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
import io, csv

from app.database import get_db
from app.models.sale import Sale
from app.models.purchase import PurchaseOrder, PurchaseItem, Supplier
from app.models.expense import Expense
from app.models.hr import Employee, SalaryPayment, Brand
from app.models.cash import CashBalance
from app.models.branch import Branch
from app.utils.auth import get_current_user
from app.models.user import User
from app.routes.hr import SALARY_VISIBLE_ROLES, _brand_branch_ids

router = APIRouter(prefix="/api/export", tags=["export"])

channels = ["cash", "knet", "link", "talabat", "keeta", "jahez"]


def _branch_map(db: Session):
    return {b.id: b.name for b in db.query(Branch).all()}


def _branch_map_ar(db: Session):
    return {b.id: (b.name_ar or b.name) for b in db.query(Branch).all()}


# ── helpers for CSV / Excel / PDF ────────────────────────────────────

def _csv_response(header: List[str], data: List[list], filename: str):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    for row in data:
        writer.writerow(row)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
    )


def _excel_response(header: List[str], data: List[list], filename: str, summary_rows: int = 0):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = filename

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Summary row styles
    total_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    payable_font = Font(bold=True, color="2E7D32", size=11)
    payable_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    hold_font = Font(bold=True, color="C62828", size=11)
    hold_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for col_idx, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Style summary rows
    if summary_rows > 0 and len(data) >= summary_rows:
        total_row_idx = len(data) - summary_rows + 2  # +2 for header row + 1-based
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
        if summary_rows >= 2:
            payable_row_idx = total_row_idx + 1
            for col_idx in range(1, len(header) + 1):
                cell = ws.cell(row=payable_row_idx, column=col_idx)
                cell.font = payable_font
                cell.fill = payable_fill
        if summary_rows >= 3:
            hold_row_idx = total_row_idx + 2
            for col_idx in range(1, len(header) + 1):
                cell = ws.cell(row=hold_row_idx, column=col_idx)
                cell.font = hold_font
                cell.fill = hold_fill

    # Auto-width
    for col_idx, h in enumerate(header, 1):
        max_len = len(str(h))
        for row in data:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


def _pdf_response(header: List[str], data: List[list], filename: str, title: str = "", summary_rows: int = 0):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    buf = io.BytesIO()
    page_size = landscape(A4) if len(header) > 8 else A4
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)

    styles = getSampleStyleSheet()
    elements = []

    if title:
        title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=14, spaceAfter=10)
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 5 * mm))

    # Truncate long strings for PDF
    def trunc(val, max_len=25):
        s = str(val) if val is not None else ""
        return s[:max_len] + ".." if len(s) > max_len else s

    table_data = [header]
    for row in data:
        table_data.append([trunc(v) for v in row])

    if not table_data or len(table_data) < 2:
        table_data.append(["No data"] + [""] * (len(header) - 1))

    num_cols = len(header)
    avail_width = page_size[0] - 20 * mm
    col_width = avail_width / num_cols

    t = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    # Highlight summary rows at the bottom
    if summary_rows > 0 and len(table_data) > summary_rows:
        total_row = len(table_data) - summary_rows  # first summary row (TOTAL)
        # TOTAL row styling
        style_cmds.append(("BACKGROUND", (0, total_row), (-1, total_row), colors.HexColor("#1B5E20")))
        style_cmds.append(("TEXTCOLOR", (0, total_row), (-1, total_row), colors.white))
        style_cmds.append(("FONTSIZE", (0, total_row), (-1, total_row), 8))
        # PAYABLE row styling
        if summary_rows >= 2:
            payable_row = total_row + 1
            style_cmds.append(("BACKGROUND", (0, payable_row), (-1, payable_row), colors.HexColor("#E8F5E9")))
            style_cmds.append(("TEXTCOLOR", (0, payable_row), (-1, payable_row), colors.HexColor("#2E7D32")))
            style_cmds.append(("FONTSIZE", (0, payable_row), (-1, payable_row), 8))
        # ON HOLD row styling
        if summary_rows >= 3:
            hold_row = total_row + 2
            style_cmds.append(("BACKGROUND", (0, hold_row), (-1, hold_row), colors.HexColor("#FFEBEE")))
            style_cmds.append(("TEXTCOLOR", (0, hold_row), (-1, hold_row), colors.HexColor("#C62828")))
            style_cmds.append(("FONTSIZE", (0, hold_row), (-1, hold_row), 8))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
    )


def _respond(fmt: str, header: List[str], data: List[list], filename: str, title: str = "", summary_rows: int = 0):
    if fmt == "excel":
        return _excel_response(header, data, filename, summary_rows=summary_rows)
    elif fmt == "pdf":
        return _pdf_response(header, data, filename, title, summary_rows=summary_rows)
    return _csv_response(header, data, filename)


# ── Data extraction helpers ─────────────────────────────────────────

def _sales_data(db, user, branch_id, brand_id=None):
    bmap = _branch_map(db)
    q = db.query(Sale)
    bb_ids = _brand_branch_ids(db, brand_id)
    if branch_id:
        q = q.filter(Sale.branch_id == branch_id)
    elif bb_ids is not None:
        q = q.filter(Sale.branch_id.in_(bb_ids))
    elif user.role == "staff" and user.branch_id:
        q = q.filter(Sale.branch_id == user.branch_id)
    rows = q.order_by(Sale.date.desc()).all()

    header = ["Date", "Branch"]
    for prefix in ["Foodics", "Physical"]:
        for ch in channels:
            header.append(f"{prefix} {ch.title()}")
        header.append(f"{prefix} Total")
    header.append("Difference")

    data = []
    for r in rows:
        row = [str(r.date), bmap.get(r.branch_id, "")]
        for prefix in ["foodics", "physical"]:
            total = 0
            for ch in channels:
                val = getattr(r, f"{prefix}_{ch}", 0) or 0
                row.append(val)
                total += val
            row.append(total)
        f_total = sum(getattr(r, f"foodics_{ch}", 0) or 0 for ch in channels)
        p_total = sum(getattr(r, f"physical_{ch}", 0) or 0 for ch in channels)
        row.append(p_total - f_total)
        data.append(row)
    return header, data


def _purchases_data(db, user, branch_id, brand_id=None):
    bmap = _branch_map(db)
    q = db.query(PurchaseOrder)
    bb_ids = _brand_branch_ids(db, brand_id)
    if branch_id:
        q = q.filter(PurchaseOrder.branch_id == branch_id)
    elif bb_ids is not None:
        q = q.filter(PurchaseOrder.branch_id.in_(bb_ids))
    elif user.role == "staff" and user.branch_id:
        q = q.filter(PurchaseOrder.branch_id == user.branch_id)
    rows = q.order_by(PurchaseOrder.date.desc()).all()
    header = ["Date", "Branch", "Supplier ID", "Payment Type", "Total Amount", "Status", "Notes"]
    data = [[str(r.date), bmap.get(r.branch_id, ""), r.supplier_id,
             r.payment_type, r.total_amount, r.status, r.notes or ""] for r in rows]
    return header, data


def _expenses_data(db, user, branch_id, brand_id=None):
    bmap = _branch_map(db)
    q = db.query(Expense)
    bb_ids = _brand_branch_ids(db, brand_id)
    if branch_id:
        q = q.filter(Expense.branch_id == branch_id)
    elif bb_ids is not None:
        q = q.filter(Expense.branch_id.in_(bb_ids))
    elif user.role == "staff" and user.branch_id:
        q = q.filter(Expense.branch_id == user.branch_id)
    rows = q.order_by(Expense.date.desc()).all()
    header = ["Date", "Branch", "Description", "Amount", "Payment Method", "Notes"]
    data = [[str(r.date), bmap.get(r.branch_id, ""), r.description,
             r.amount, r.payment_method, r.notes or ""] for r in rows]
    return header, data


def _hr_data(db, user, branch_id, brand_id=None):
    bmap = _branch_map(db)
    q = db.query(Employee)
    bb_ids = _brand_branch_ids(db, brand_id)
    if branch_id:
        q = q.filter(Employee.branch_id == branch_id)
    elif bb_ids is not None:
        q = q.filter(Employee.branch_id.in_(bb_ids))
    elif user.role == "staff" and user.branch_id:
        q = q.filter(Employee.branch_id == user.branch_id)
    rows = q.order_by(Employee.name).all()
    header = ["Staff No.", "Name", "Name (AR)", "Branch", "Civil ID", "Position", "Phone",
              "Employer", "Work Permit Salary", "Actual Salary", "Salary Transfer", "IBAN", "Bank",
              "Join Date", "Active"]
    data = [[r.staff_no or "", r.name, r.name_ar or "", bmap.get(r.branch_id, ""), r.civil_id or "",
             r.position or "", r.phone or "", r.employer or "",
             r.work_permit_salary or 0, r.actual_salary or 0,
             r.salary_transfer_method or "", r.iban or "", r.bank_name or "",
             str(r.join_date) if r.join_date else "", "Yes" if r.is_active else "No"] for r in rows]
    return header, data


def _cash_data(db, user, branch_id, brand_id=None):
    bmap = _branch_map(db)
    q = db.query(CashBalance)
    bb_ids = _brand_branch_ids(db, brand_id)
    if branch_id:
        q = q.filter(CashBalance.branch_id == branch_id)
    elif bb_ids is not None:
        q = q.filter(CashBalance.branch_id.in_(bb_ids))
    rows = q.order_by(CashBalance.date.desc()).all()
    header = ["Date", "Branch", "Opening Balance", "Cash Sales", "Petty Cash In",
              "Cash Purchases", "Cash Expenses", "Cash Withdrawn", "Deposited", "Closing Balance"]
    data = [[str(r.date), bmap.get(r.branch_id, ""), r.opening_balance, r.cash_sales,
             r.petty_cash_in, r.cash_purchases, r.cash_expenses, r.cash_withdrawn,
             r.deposited, r.closing_balance] for r in rows]
    return header, data


def _salary_data(db, user, month, lang: str = "en", brand_id=None):
    is_ar = lang == "ar"
    bmap = _branch_map_ar(db) if is_ar else _branch_map(db)
    emp_map = {e.id: e for e in db.query(Employee).all()}
    q = db.query(SalaryPayment)
    if month:
        q = q.filter(SalaryPayment.month == month)
    bb_ids = _brand_branch_ids(db, brand_id)
    if bb_ids is not None:
        q = q.filter(SalaryPayment.branch_id.in_(bb_ids))
    rows = q.order_by(SalaryPayment.month.desc()).all()
    if is_ar:
        header = ["الشهر", "رقم الموظف", "الاسم", "المنصب", "الفرع", "أيام العمل",
                  "الراتب الأساسي", "حافز", "مكافأة", "راتب الإجازة", "تذكرة", "عمل إضافي",
                  "إجمالي البدلات", "خصم الغياب", "خصم القرض", "الغرامة",
                  "خصم آخر", "إجمالي الخصومات", "صافي الراتب", "طريقة الدفع", "الحالة"]
    else:
        header = ["Month", "Staff No.", "Name", "Position", "Branch", "Days Worked",
                  "Basic Salary", "Incentive", "Bonus", "Leave Salary", "Ticket", "Overtime",
                  "Total Allowances", "Absence Deduction", "Loan Deduction", "Penalty",
                  "Other Deduction", "Total Deductions", "Net Salary", "Payment Method", "Status"]
    data = []
    # Accumulators for totals
    sum_basic = 0
    sum_incentive = 0
    sum_bonus = 0
    sum_leave_salary = 0
    sum_ticket = 0
    sum_overtime = 0
    sum_allowances = 0
    sum_absence_ded = 0
    sum_loan_ded = 0
    sum_penalty = 0
    sum_other_ded = 0
    sum_deductions = 0
    sum_net = 0
    sum_on_hold = 0
    sum_payable = 0
    # Filter out records for employees with 0 or null actual salary
    rows = [r for r in rows if emp_map.get(r.employee_id) and (emp_map[r.employee_id].actual_salary or 0) > 0]

    for r in rows:
        emp = emp_map.get(r.employee_id)
        incentive = r.incentive or 0
        bonus = r.bonus or 0
        leave_sal = r.leave_salary or 0
        ticket = r.ticket_payment or 0
        overtime = r.overtime or 0
        allowances = r.allowances or 0
        absence_ded = r.absence_deduction or 0
        loan_ded = r.loan_deduction or 0
        penalty = r.penalty or 0
        other_ded = r.other_deduction or 0
        deductions = r.deductions or 0
        net = r.net_salary or 0
        emp_name = ""
        if emp:
            emp_name = (emp.name_ar or emp.name) if is_ar else emp.name
        data.append([
            r.month, emp.staff_no if emp else "", emp_name,
            emp.position if emp else "", bmap.get(emp.branch_id, "") if emp else "",
            r.days_worked or 30, r.basic_salary,
            incentive, bonus, leave_sal, ticket, overtime,
            allowances, absence_ded, loan_ded, penalty,
            other_ded, deductions, net,
            r.payment_method or "", r.status,
        ])
        sum_basic += r.basic_salary or 0
        sum_incentive += incentive
        sum_bonus += bonus
        sum_leave_salary += leave_sal
        sum_ticket += ticket
        sum_overtime += overtime
        sum_allowances += allowances
        sum_absence_ded += absence_ded
        sum_loan_ded += loan_ded
        sum_penalty += penalty
        sum_other_ded += other_ded
        sum_deductions += deductions
        sum_net += net
        if r.status == "on_hold":
            sum_on_hold += net
        elif r.status == "pending":
            sum_payable += net

    # Append totals row
    total_label = "الإجمالي" if is_ar else "TOTAL / الإجمالي"
    payable_label = "المستحق الدفع" if is_ar else "PAYABLE / المستحق الدفع"
    hold_label = "معلق" if is_ar else "ON HOLD / معلق"
    if data:
        data.append([
            "", "", total_label, "", "", "",
            round(sum_basic, 3),
            round(sum_incentive, 3), round(sum_bonus, 3),
            round(sum_leave_salary, 3), round(sum_ticket, 3), round(sum_overtime, 3),
            round(sum_allowances, 3),
            round(sum_absence_ded, 3), round(sum_loan_ded, 3), round(sum_penalty, 3),
            round(sum_other_ded, 3), round(sum_deductions, 3), round(sum_net, 3),
            "", "",
        ])
        # Append on-hold and payable summary rows
        data.append([
            "", "", payable_label, "", "", "",
            "", "", "", "", "", "",
            "", "", "", "", "", "", round(sum_payable, 3),
            "", "pending",
        ])
        data.append([
            "", "", hold_label, "", "", "",
            "", "", "", "", "", "",
            "", "", "", "", "", "", round(sum_on_hold, 3),
            "", "on_hold",
        ])
    return header, data


# ── Unified endpoints: /api/export/{module}/{format} ────────────────

@router.get("/sales/{fmt}")
def export_sales(fmt: str, branch_id: Optional[int] = None, brand_id: Optional[int] = None,
                 db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    header, data = _sales_data(db, user, branch_id, brand_id=brand_id)
    return _respond(fmt, header, data, "sales", "Sales Report")


@router.get("/purchases/{fmt}")
def export_purchases(fmt: str, branch_id: Optional[int] = None, brand_id: Optional[int] = None,
                     db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    header, data = _purchases_data(db, user, branch_id, brand_id=brand_id)
    return _respond(fmt, header, data, "purchases", "Purchases Report")


@router.get("/expenses/{fmt}")
def export_expenses(fmt: str, branch_id: Optional[int] = None, brand_id: Optional[int] = None,
                    db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    header, data = _expenses_data(db, user, branch_id, brand_id=brand_id)
    return _respond(fmt, header, data, "expenses", "Expenses Report")


@router.get("/hr/{fmt}")
def export_hr(fmt: str, branch_id: Optional[int] = None, brand_id: Optional[int] = None,
              db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    header, data = _hr_data(db, user, branch_id, brand_id=brand_id)
    return _respond(fmt, header, data, "employees", "Employees Report")


@router.get("/cash/{fmt}")
def export_cash(fmt: str, branch_id: Optional[int] = None, brand_id: Optional[int] = None,
                db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    header, data = _cash_data(db, user, branch_id, brand_id=brand_id)
    return _respond(fmt, header, data, "cash_management", "Cash Management Report")


# ── Bulk Payslips PDF (one full A4 page per employee) ───────────────
# NOTE: This route MUST be defined before /salary/{fmt} to avoid being
# swallowed by the path-parameter route.

@router.get("/salary/slips/pdf")
def export_salary_slips_pdf(
    month: Optional[str] = None,
    lang: Optional[str] = None,
    brand_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if user.role not in SALARY_VISIBLE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import arabic_reshaper
    from bidi.algorithm import get_display

    # Register DejaVuSans – supports Arabic glyphs
    _dvs = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    _dvsb = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans", _dvs))
    if "DejaVuSans-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", _dvsb))

    def _ar(text: str) -> str:
        """Reshape and reorder Arabic text for PDF rendering."""
        if not text:
            return text
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)

    is_ar = (lang or "en") == "ar"
    bmap = _branch_map_ar(db) if is_ar else _branch_map(db)
    emp_map = {e.id: e for e in db.query(Employee).all()}

    q = db.query(SalaryPayment)
    if month:
        q = q.filter(SalaryPayment.month == month)
    bb_ids = _brand_branch_ids(db, brand_id)
    if bb_ids is not None:
        q = q.filter(SalaryPayment.branch_id.in_(bb_ids))
    records = q.all()
    # Filter out employees with 0 actual salary
    records = [r for r in records if emp_map.get(r.employee_id) and (emp_map[r.employee_id].actual_salary or 0) > 0]
    # Sort by branch then staff_no within each branch
    records.sort(key=lambda r: (
        bmap.get(emp_map[r.employee_id].branch_id, "") or "",
        emp_map[r.employee_id].staff_no or "",
    ))

    if not records:
        raise HTTPException(404, "No salary records found")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)

    # Styles – use DejaVuSans for Arabic glyph support
    FONT = "DejaVuSans"
    FONTB = "DejaVuSans-Bold"
    hdr_style = ParagraphStyle("hdr", fontSize=16, alignment=1, spaceAfter=2,
                                fontName=FONTB)
    hdr_ar_style = ParagraphStyle("hdr_ar", fontSize=14, alignment=1, spaceAfter=2,
                                   fontName=FONT)
    sub_style = ParagraphStyle("sub", fontSize=12, alignment=1, spaceAfter=6,
                                fontName=FONTB, textColor=colors.HexColor("#555"))
    month_style = ParagraphStyle("month", fontSize=10, alignment=1, spaceAfter=8,
                                  fontName=FONT)
    sect_style = ParagraphStyle("sect", fontSize=10, fontName=FONTB,
                                 backColor=colors.HexColor("#E8F5E9"),
                                 borderPadding=(4, 6, 4, 6), spaceAfter=4)
    lbl = ParagraphStyle("lbl", fontSize=9, fontName=FONT)
    val = ParagraphStyle("val", fontSize=9, fontName=FONTB, alignment=2)
    net_lbl = ParagraphStyle("net_lbl", fontSize=12, fontName=FONTB,
                              textColor=colors.HexColor("#1B5E20"))
    net_val = ParagraphStyle("net_val", fontSize=12, fontName=FONTB,
                              alignment=2, textColor=colors.HexColor("#1B5E20"))
    green_val = ParagraphStyle("green_val", fontSize=9, fontName=FONTB,
                                alignment=2, textColor=colors.HexColor("#2E7D32"))
    red_val = ParagraphStyle("red_val", fontSize=9, fontName=FONTB,
                              alignment=2, textColor=colors.HexColor("#C62828"))
    sig_style = ParagraphStyle("sig", fontSize=8, alignment=1, spaceBefore=4,
                                fontName=FONT)

    elements = []

    for idx, sp in enumerate(records):
        emp = emp_map.get(sp.employee_id)
        if not emp:
            continue
        branch_name = bmap.get(emp.branch_id, "")
        emp_name = (emp.name_ar or emp.name) if is_ar else emp.name

        # ── Header ──
        elements.append(Paragraph("WAHID MUDAWWARAH RESTAURANT", hdr_style))
        elements.append(Paragraph(_ar("مطعم واحد مدوّرة"), hdr_ar_style))
        elements.append(Paragraph(f"PAY SLIP / {_ar('قسيمة الراتب')}", sub_style))
        month_label = sp.month or ""
        if is_ar:
            elements.append(Paragraph(f"{_ar('الشهر')}: <b>{month_label}</b>", month_style))
        else:
            elements.append(Paragraph(f"Month: <b>{month_label}</b>", month_style))
        elements.append(Spacer(1, 4 * mm))

        # ── Employee Info ──
        elements.append(Paragraph(f"Employee Information / {_ar('معلومات الموظف')}", sect_style))
        _name_display = _ar(emp_name) if emp_name else "—"
        _pos_display = _ar(emp.position) if emp.position else "—"
        _branch_display = _ar(branch_name) if branch_name else "—"
        _bank_display = _ar(emp.bank_name) if emp.bank_name else "—"
        emp_info = [
            [Paragraph(f"Staff No. / {_ar('رقم الموظف')}", lbl), Paragraph(emp.staff_no or "—", val),
             Paragraph(f"Name / {_ar('الاسم')}", lbl), Paragraph(_name_display, val)],
            [Paragraph(f"Position / {_ar('المسمى الوظيفي')}", lbl), Paragraph(_pos_display, val),
             Paragraph(f"Branch / {_ar('الفرع')}", lbl), Paragraph(_branch_display, val)],
            [Paragraph(f"Civil ID / {_ar('الرقم المدني')}", lbl), Paragraph(emp.civil_id or "—", val),
             Paragraph("IBAN", lbl), Paragraph(emp.iban or "—", val)],
            [Paragraph(f"Bank / {_ar('البنك')}", lbl), Paragraph(_bank_display, val),
             Paragraph(f"Join Date / {_ar('تاريخ الالتحاق')}", lbl),
             Paragraph(str(emp.join_date) if emp.join_date else "—", val)],
        ]
        avail = A4[0] - 30 * mm
        emp_t = Table(emp_info, colWidths=[avail * 0.22, avail * 0.28, avail * 0.22, avail * 0.28])
        emp_t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#FAFAFA")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#FAFAFA")),
        ]))
        elements.append(emp_t)
        elements.append(Spacer(1, 4 * mm))

        # ── Salary Details ──
        elements.append(Paragraph(f"Salary Details / {_ar('تفاصيل الراتب')}", sect_style))

        def _fmt(v):
            return f"{v:.3f}"

        salary_rows = [
            [Paragraph(f"Basic Salary / {_ar('الراتب الأساسي')}", lbl),
             Paragraph(f"KD {_fmt(sp.basic_salary)}", val)],
            [Paragraph(f"Days Worked / {_ar('أيام العمل')}", lbl),
             Paragraph(f"{sp.days_worked or 30} / {sp.total_days or 30}", val)],
            [Paragraph(f"Period / {_ar('الفترة')}", lbl),
             Paragraph(f"{sp.period_start or '—'}  to  {sp.period_end or '—'}", val)],
        ]
        sal_t = Table(salary_rows, colWidths=[avail * 0.6, avail * 0.4])
        sal_t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(sal_t)
        elements.append(Spacer(1, 3 * mm))

        # ── Earnings ──
        elements.append(Paragraph(f"Earnings / {_ar('المستحقات')}", sect_style))
        earn_rows = []
        earn_items = [
            (f"Overtime / {_ar('العمل الإضافي')}", sp.overtime or 0),
            (f"Bonus / {_ar('مكافأة')}", sp.bonus or 0),
            (f"Incentive / {_ar('حافز')}", sp.incentive or 0),
            (f"Leave Salary / {_ar('راتب الإجازة')}", sp.leave_salary or 0),
            (f"Ticket Payment / {_ar('تذكرة السفر')}", sp.ticket_payment or 0),
            (f"Housing Allowance / {_ar('بدل سكن')}", sp.housing_allowance or 0),
            (f"Transport Allowance / {_ar('بدل نقل')}", sp.transport_allowance or 0),
            (f"Food Allowance / {_ar('بدل طعام')}", sp.food_allowance or 0),
            (f"Other Allowance / {_ar('بدلات أخرى')}", sp.other_allowance or 0),
        ]
        for label_text, amount in earn_items:
            if amount > 0:
                earn_rows.append([
                    Paragraph(label_text, lbl),
                    Paragraph(f"+{_fmt(amount)}", green_val),
                ])
        total_earnings = sp.allowances or 0
        earn_rows.append([
            Paragraph(f"<b>Total Earnings / {_ar('إجمالي المستحقات')}</b>", lbl),
            Paragraph(f"<b>KD {_fmt(total_earnings)}</b>", green_val),
        ])
        earn_t = Table(earn_rows, colWidths=[avail * 0.6, avail * 0.4])
        earn_style = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
        ]
        earn_t.setStyle(TableStyle(earn_style))
        elements.append(earn_t)
        elements.append(Spacer(1, 3 * mm))

        # ── Deductions ──
        elements.append(Paragraph(f"Deductions / {_ar('الخصومات')}", sect_style))
        ded_rows = []
        ded_items = [
            (f"Absence Deduction / {_ar('خصم غياب')}", sp.absence_deduction or 0),
            (f"Loan Deduction / {_ar('خصم القرض')}", sp.loan_deduction or 0),
            (f"Penalty/Fine / {_ar('غرامة')}", sp.penalty or 0),
            (f"Late Deduction / {_ar('خصم تأخير')}", sp.late_deduction or 0),
            (f"Other Deduction / {_ar('خصومات أخرى')}", sp.other_deduction or 0),
            (f"Advance / {_ar('سلفة')}", sp.advance or 0),
        ]
        for label_text, amount in ded_items:
            if amount > 0:
                ded_rows.append([
                    Paragraph(label_text, lbl),
                    Paragraph(f"-{_fmt(amount)}", red_val),
                ])
        total_deductions = sp.deductions or 0
        ded_rows.append([
            Paragraph(f"<b>Total Deductions / {_ar('إجمالي الخصومات')}</b>", lbl),
            Paragraph(f"<b>KD {_fmt(total_deductions)}</b>", red_val),
        ])
        ded_t = Table(ded_rows, colWidths=[avail * 0.6, avail * 0.4])
        ded_style = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFEBEE")),
        ]
        ded_t.setStyle(TableStyle(ded_style))
        elements.append(ded_t)
        elements.append(Spacer(1, 4 * mm))

        # ── Net Salary ──
        net_rows = [
            [Paragraph(f"NET SALARY / {_ar('صافي الراتب')}", net_lbl),
             Paragraph(f"KD {_fmt(sp.net_salary or 0)}", net_val)],
            [Paragraph(f"Payment Method / {_ar('طريقة الدفع')}", lbl),
             Paragraph(f"Bank Transfer / {_ar('تحويل بنكي')}" if sp.payment_method == "bank_transfer" else f"Cash / {_ar('نقداً')}", val)],
            [Paragraph(f"Status / {_ar('الحالة')}", lbl),
             Paragraph(sp.status.upper() if sp.status else "PENDING", val)],
        ]
        if sp.status == "paid" and sp.paid_date:
            net_rows.append([
                Paragraph(f"Paid Date / {_ar('تاريخ الدفع')}", lbl),
                Paragraph(str(sp.paid_date), val),
            ])
        net_t = Table(net_rows, colWidths=[avail * 0.6, avail * 0.4])
        net_t.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#1B5E20")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(net_t)
        elements.append(Spacer(1, 10 * mm))

        # ── Signatures ──
        sig_rows = [[
            Paragraph(f"____________________________<br/>Employee Signature<br/>{_ar('توقيع الموظف')}", sig_style),
            Paragraph("", sig_style),
            Paragraph(f"____________________________<br/>Authorized Signature<br/>{_ar('التوقيع المعتمد')}", sig_style),
        ]]
        sig_t = Table(sig_rows, colWidths=[avail * 0.4, avail * 0.2, avail * 0.4])
        sig_t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
            ("TOPPADDING", (0, 0), (-1, -1), 15),
        ]))
        elements.append(sig_t)

        # Page break between slips (not after the last one)
        if idx < len(records) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    fname = f"payslips_{month or 'all'}"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}.pdf"},
    )


@router.get("/salary/{fmt}")
def export_salary(fmt: str, month: Optional[str] = None, lang: Optional[str] = None,
                  brand_id: Optional[int] = None,
                  db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if user.role not in SALARY_VISIBLE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")
    language = lang or "en"
    header, data = _salary_data(db, user, month, lang=language, brand_id=brand_id)
    title = f"كشف الرواتب - {month or 'الكل'}" if language == "ar" else f"Salary Sheet - {month or 'All'}"
    return _respond(fmt, header, data, f"salary_{month or 'all'}", title, summary_rows=3)


# ── Purchase Order PDF ───────────────────────────────────────────────


@router.get("/purchase-order/{order_id}/pdf")
def export_purchase_order_pdf(order_id: int, db: Session = Depends(get_db),
                              _=Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == order_id).first()
    if not po:
        raise HTTPException(404, "Order not found")

    supplier = db.query(Supplier).filter(Supplier.id == po.supplier_id).first()
    branch = db.query(Branch).filter(Branch.id == po.branch_id).first()
    items = db.query(PurchaseItem).filter(PurchaseItem.purchase_order_id == order_id).all()

    # Get brand from branch
    brand = None
    if branch and branch.brand_id:
        brand = db.query(Brand).filter(Brand.id == branch.brand_id).first()

    # Register font for Arabic support
    _dvs = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    _dvsb = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    if os.path.isfile(_dvs) and "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans", _dvs))
    if os.path.isfile(_dvsb) and "DejaVuSans-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", _dvsb))

    font_name = "DejaVuSans" if "DejaVuSans" in pdfmetrics.getRegisteredFontNames() else "Helvetica"
    font_bold = "DejaVuSans-Bold" if "DejaVuSans-Bold" in pdfmetrics.getRegisteredFontNames() else "Helvetica-Bold"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)

    elements = []

    # Styles
    title_style = ParagraphStyle("po_title", fontName=font_bold, fontSize=18,
                                  alignment=1, spaceAfter=2 * mm)
    subtitle_style = ParagraphStyle("po_subtitle", fontName=font_name, fontSize=10,
                                     alignment=1, textColor=colors.grey, spaceAfter=5 * mm)
    heading_style = ParagraphStyle("po_heading", fontName=font_bold, fontSize=11,
                                    spaceAfter=2 * mm)
    normal_style = ParagraphStyle("po_normal", fontName=font_name, fontSize=9,
                                   spaceAfter=1 * mm)

    # Header
    company_name = brand.name_en if brand else "Mudawwarah"
    elements.append(Paragraph(company_name, title_style))
    elements.append(Paragraph("PURCHASE ORDER", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2E7D32")))
    elements.append(Spacer(1, 5 * mm))

    # PO Info and Supplier Info side by side
    po_info = [
        [Paragraph(f"<b>PO Number:</b> PO-{po.id:04d}", normal_style),
         Paragraph(f"<b>Date:</b> {po.date}", normal_style)],
        [Paragraph(f"<b>Branch:</b> {branch.name if branch else 'N/A'}", normal_style),
         Paragraph(f"<b>Payment:</b> {po.payment_type.title()}", normal_style)],
        [Paragraph(f"<b>Status:</b> {po.status.title()}", normal_style),
         Paragraph(f"<b>Delivery:</b> {po.delivery_location or 'N/A'}", normal_style)],
    ]
    info_table = Table(po_info, colWidths=[90 * mm, 90 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 4 * mm))

    # Supplier box
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph("<b>Supplier:</b>", heading_style))
    supp_name = supplier.name if supplier else "N/A"
    supp_whatsapp = supplier.whatsapp if supplier and supplier.whatsapp else "N/A"
    elements.append(Paragraph(f"Name: {supp_name}", normal_style))
    elements.append(Paragraph(f"WhatsApp: {supp_whatsapp}", normal_style))
    elements.append(Spacer(1, 5 * mm))

    # Items table
    elements.append(Paragraph("<b>Order Items:</b>", heading_style))
    elements.append(Spacer(1, 2 * mm))

    table_header = ["#", "Item Name", "Quantity", "Unit", "Unit Price (KD)", "Total (KD)"]
    table_data = [table_header]
    grand_total = 0.0
    for idx, item in enumerate(items, 1):
        row_total = item.total or (item.quantity * item.unit_price)
        grand_total += row_total
        table_data.append([
            str(idx),
            item.item_name,
            f"{item.quantity:.2f}",
            item.unit,
            f"{item.unit_price:.3f}",
            f"{row_total:.3f}",
        ])

    # Grand total row
    table_data.append(["", "", "", "", "TOTAL:", f"KD {grand_total:.3f}"])

    col_widths = [12 * mm, 60 * mm, 25 * mm, 20 * mm, 30 * mm, 30 * mm]
    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), font_bold),
        ("FONTNAME", (0, 1), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (-2, 0), (-2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -2), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F5F5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        # Total row
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
        ("FONTNAME", (0, -1), (-1, -1), font_bold),
        ("FONTSIZE", (0, -1), (-1, -1), 10),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#2E7D32")),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 8 * mm))

    # Notes
    if po.notes:
        elements.append(Paragraph(f"<b>Notes:</b> {po.notes}", normal_style))
        elements.append(Spacer(1, 5 * mm))

    # Signature lines
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    elements.append(Spacer(1, 10 * mm))
    sig_data = [
        [Paragraph("<b>Prepared By:</b>", normal_style),
         Paragraph("<b>Approved By:</b>", normal_style),
         Paragraph("<b>Received By:</b>", normal_style)],
        ["_________________", "_________________", "_________________"],
    ]
    sig_table = Table(sig_data, colWidths=[60 * mm, 60 * mm, 60 * mm])
    sig_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buf.seek(0)
    fname = f"PO-{po.id:04d}_{supplier.name if supplier else 'order'}"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}.pdf"},
    )
